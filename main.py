"""
title           :main.py
description     :This will convert graph (naver keyword result ) to text or csv file
author          :DevHyung
date            :2018.11.04
version         :1.0.0
usage           :python3 main.py
python_version  :3.6
required module :selenium+chromewebdriver, csv, bs4
"""
from selenium import webdriver
from selenium.webdriver import ActionChains
from bs4 import BeautifulSoup
import time
from openpyxl import load_workbook, Workbook
import os
import re
pattern = re.compile(r'\s+')

def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        for depth1List in _DATA:
            sheet.append(depth1List)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = 'result'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        book.save(_FILENAME)
FILENAME = "keyword.xlsx"
HEADER = ['키워드', 'PC', 'MOBILE']
save_excel(FILENAME,None,HEADER)

if __name__=="__main__":
    NAVER_ID = input("Naver ID 입력 :: ")
    NAVER_PW = input("Naver PW 입력 :: ")
    print(">>> Made By Pakr HyungJune copyright @ DevHyung")
    # Setting variable
    print("_" * 30)
    print(">>> Loding....")
    dir = './chromedriver'  # Driver Path
    driver = webdriver.Chrome(dir)
    driver.maximize_window()
    # Login start
    driver.get("https://searchad.naver.com/")  # target page
    driver.find_element_by_xpath('//*[@id="uid"]').send_keys(NAVER_ID)
    driver.find_element_by_xpath('//*[@id="upw"]').send_keys(NAVER_PW)
    driver.find_element_by_xpath('//*[@id="container"]/main/div/div[1]/home-login/div/fieldset/span/button').click()
    time.sleep(1)
    driver.find_element_by_xpath('/html/body/my-app/wrap/welcome-beginner-layer-popup/div[2]/div[1]/a').click()
    driver.find_element_by_xpath('//*[@id="container"]/my-screen/div/div[1]/div/my-screen-board/div/div[1]/ul/li[3]/a').click()
    time.sleep(1)
    # 이부분을 수정 하면 됌 내꺼는 2
    driver.switch_to.window(driver.window_handles[1])
    while True:
        f = open("keyword.txt", 'r', encoding='utf8')
        keywordList = f.readlines()
        print(">>> {} 개 키워드 검색 시작".format(len(keywordList)))
        # if event evoke, parsing start
        idx = 1
        dataList = []
        for keyword in keywordList:
            k = re.sub(pattern, '', keyword)
            while True:
                try:
                    driver.find_element_by_xpath('//*[@id="wrap"]/div/div/div[1]/div[1]/div/div/div/div[2]/div[1]/div[1]/div[2]/form/div[1]/div/div/textarea').clear()
                    break
                except:
                    time.sleep(0.3)
            driver.find_element_by_xpath('//*[@id="wrap"]/div/div/div[1]/div[1]/div/div/div/div[2]/div[1]/div[1]/div[2]/form/div[1]/div/div/textarea').send_keys(k)
            driver.find_element_by_xpath('//*[@id="wrap"]/div/div/div[1]/div[1]/div/div/div/div[2]/div[1]/div[1]/div[2]/form/div[4]/div/div/ul/li/button').click()
            time.sleep(0.3)
            print(">>> {}번째, {} 진행중..".format(idx,k))
            idx += 1
            loopIdx = 1
            while True:
                try:
                    isSearch = False
                    bs4 = BeautifulSoup(driver.page_source, "lxml")
                    tr = bs4.find('table',class_='table table-bordered').find_all('tr')
                    tmp = tr[2]
                    if k == tmp['row-id'].strip().replace(' ',''):#같은경우
                        try:
                            pc = tmp.find_all('td',class_=' text-right txt-r')[0].get_text().replace(',','')
                        except:
                            pc = "ERROR"
                        try:
                            mobile = tmp.find_all('td', class_=' text-right txt-r')[1].get_text().replace(',','')
                        except:
                            mobile = "ERROR"
                        dataList.append([k,pc,mobile])
                        if len(dataList) > 5:
                            save_excel(FILENAME, dataList, None)
                            dataList.clear()
                        print("\t>>> OK".format(k))
                        isSearch = True


                    if isSearch:
                        break
                    else:
                        time.sleep(0.5)
                        loopIdx += 1
                        if loopIdx == 3:
                            dataList.append([k, '확인 바람', '확인 바람'])
                            print("\t>>> {} 확인필요 ..".format(k))
                            break
                except:
                    time.sleep(0.3)

        save_excel(FILENAME, dataList, None)

        inputNum = input(">>> 끝내려면 0을 입력하세요 아니면 파일을 바꾸고 엔터를 누르세요")
        if inputNum == '0':
            break
